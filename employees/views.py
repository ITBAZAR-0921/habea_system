from datetime import date, datetime

from django.contrib import messages
from django.contrib.auth.models import User
from django.shortcuts import redirect, render
from django.utils import timezone
from django.utils.dateparse import parse_date
from django.utils.text import slugify

from config.permissions import role_required

from .forms import EmployeeExcelUploadForm
from .models import Department, Employee


def _normalize_header(value):
    if value is None:
        return ''
    return str(value).strip().lower().replace(' ', '').replace('_', '').replace('-', '')


def _as_text(value):
    if value is None:
        return ''
    return str(value).strip()


def _to_bool(value):
    if value is None:
        return False
    normalized = str(value).strip().lower()
    return normalized in {'1', 'true', 'yes', 'y', 'тийм'}


def _to_date(value):
    if value in (None, ''):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return parse_date(str(value).strip())


def _build_username(first_name, last_name, register):
    base = slugify(f'{first_name}.{last_name}') or slugify(str(register)) or 'employee'
    username = base
    i = 1
    while User.objects.filter(username=username).exists():
        i += 1
        username = f'{base}{i}'
    return username


@role_required(['system_admin', 'hse_manager', 'department_head', 'employee'])
def employee_list(request):
    employees = Employee.objects.select_related('department').all().order_by('last_name', 'first_name')
    return render(request, 'employees/employee_list.html', {'employees': employees})


@role_required(['system_admin', 'hse_manager'])
def employee_upload(request):
    form = EmployeeExcelUploadForm(request.POST or None, request.FILES or None)
    if request.method == 'POST' and form.is_valid():
        try:
            from openpyxl import load_workbook
        except ImportError:
            messages.error(request, 'openpyxl сан суулгагдаагүй байна. requirements суулгана уу.')
            return redirect('employee_upload')

        workbook = load_workbook(form.cleaned_data['file'], data_only=True)
        sheet = workbook.active

        headers = [cell.value for cell in sheet[1]]
        header_map = {_normalize_header(name): idx for idx, name in enumerate(headers)}

        def value_from_row(row, *aliases):
            for alias in aliases:
                idx = header_map.get(alias)
                if idx is not None:
                    return row[idx]
            return None

        created_count = 0
        updated_count = 0
        skipped_count = 0
        errors = []

        for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not any(row):
                continue

            first_name = _as_text(value_from_row(row, 'firstname'))
            last_name = _as_text(value_from_row(row, 'lastname'))
            register = _as_text(value_from_row(row, 'register', 'registernumber', 'registernum', 'registerno'))
            phone = _as_text(value_from_row(row, 'phone', 'phonenumber'))
            department_name = _as_text(value_from_row(row, 'department', 'dept'))
            username = _as_text(value_from_row(row, 'username', 'login'))
            password = value_from_row(row, 'password', 'pass')
            email = _as_text(value_from_row(row, 'email', 'mail'))
            position = _as_text(value_from_row(row, 'position', 'jobtitle'))
            hired_date = _to_date(value_from_row(row, 'hireddate', 'startdate', 'employmentdate'))
            is_head = _to_bool(value_from_row(row, 'ishead', 'head', 'isdepartmenthead'))

            if not first_name or not last_name or not register:
                skipped_count += 1
                errors.append(f'{row_number}-р мөр: first_name, last_name, register_number заавал байна.')
                continue

            if not username:
                username = _build_username(first_name, last_name, register)

            user = User.objects.filter(username=username).first()
            created_user = False
            if not user:
                user = User.objects.create_user(username=username)
                created_user = True

            user.first_name = first_name
            user.last_name = last_name
            if email:
                user.email = email

            if password not in (None, ''):
                user.set_password(str(password))
            elif created_user:
                user.set_password('ChangeMe123!')
            user.save()

            department = None
            if department_name:
                department, _ = Department.objects.get_or_create(name=department_name)

            employee = Employee.objects.filter(register=register).first() or Employee.objects.filter(user=user).first()

            payload = {
                'user': user,
                'first_name': first_name,
                'last_name': last_name,
                'register': register,
                'position': position or ('Хэлтсийн дарга' if is_head else 'Ажилтан'),
                'department': department,
                'phone': phone,
                'email': email,
                'hired_date': hired_date or timezone.localdate(),
                'is_head': is_head,
            }

            if employee:
                for key, value in payload.items():
                    setattr(employee, key, value)
                employee.save()
                updated_count += 1
            else:
                Employee.objects.create(**payload)
                created_count += 1

        messages.success(
            request,
            f'Excel импорт амжилттай: шинэ {created_count}, шинэчлэгдсэн {updated_count}, алгассан {skipped_count}.',
        )
        for error in errors[:10]:
            messages.warning(request, error)

        return redirect('employee_list')

    return render(request, 'employees/employee_upload.html', {'form': form})
